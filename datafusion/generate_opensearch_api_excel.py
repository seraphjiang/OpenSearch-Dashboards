#!/usr/bin/env python3
"""
Generate Excel spreadsheet organized by OpenSearch API
showing which OSD endpoints/features use each OpenSearch API
"""

import re
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def parse_opensearch_api_usage(file_path):
    """Parse the markdown file and extract OpenSearch API usage"""
    with open(file_path, 'r', encoding='utf-8') as f:
        content = f.read()

    # Find the cross-reference section
    cross_ref_pattern = r'## Cross-Reference: OpenSearch API to OSD Endpoints(.*?)(?=\n## |\Z)'
    match = re.search(cross_ref_pattern, content, re.DOTALL)

    opensearch_apis = {}

    if match:
        cross_ref_content = match.group(1)

        # Parse each OpenSearch API section
        # Pattern: **`api_name`** or **`category.api_name`**
        api_sections = re.findall(r'\*\*`([^`]+)`\*\*\s*\n(.*?)(?=\n\*\*`|\Z)', cross_ref_content, re.DOTALL)

        for api_name, usage_text in api_sections:
            # Clean up API name
            api_name = api_name.strip()

            # Extract used by information
            used_by = []

            # Look for bullet points or lines mentioning endpoints/plugins
            lines = usage_text.strip().split('\n')
            for line in lines:
                line = line.strip()
                if line and not line.startswith('#'):
                    # Remove markdown formatting
                    clean_line = re.sub(r'\*\*|`', '', line)
                    clean_line = re.sub(r'^[-•*]\s*', '', clean_line)
                    if clean_line:
                        used_by.append(clean_line)

            opensearch_apis[api_name] = used_by

    # Also parse from main mapping tables
    sections = [
        ('Core: Saved Objects API', 'Saved Objects'),
        ('Core: Status & Health', 'Status/Health'),
        ('Plugin: Data (Search)', 'Data/Search'),
        ('Plugin: Console (Dev Tools Proxy)', 'Console'),
        ('Plugin: Data Importer', 'Data Importer'),
        ('Plugin: Application Config', 'Application Config'),
        ('Plugin: Region Map', 'Region Map'),
        ('Plugin: Query Enhancements', 'Query Enhancements'),
        ('Plugin: Data Source Management', 'Data Source Management'),
        ('Plugin: Chat (ML/AI)', 'Chat/ML'),
        ('Plugin: Workspace', 'Workspace'),
        ('Plugin: Home (Sample Data)', 'Home'),
        ('Plugin: Telemetry', 'Telemetry'),
        ('Plugin: Saved Objects Management', 'Saved Objects Mgmt'),
        ('Core: Dynamic Config Store (Internal)', 'Config Store'),
        ('Core: Saved Objects Migrations (Internal)', 'Migrations'),
    ]

    api_usage_map = {}

    for section_header, plugin_name in sections:
        pattern = rf'## {re.escape(section_header)}.*?\n(.*?)(?=\n## |\Z)'
        match = re.search(pattern, content, re.DOTALL)

        if not match:
            continue

        section_content = match.group(1)

        # Look for endpoint mapping tables
        table_pattern = r'\| OSD Endpoint \| Method \| OpenSearch APIs Called \| Data Flow \|.*?\n\|.*?\n((?:\|.*?\n)*)'
        table_match = re.search(table_pattern, section_content, re.DOTALL)

        if table_match:
            table_rows = table_match.group(1).strip().split('\n')
            for row in table_rows:
                cells = [cell.strip() for cell in row.split('|')[1:-1]]
                if len(cells) >= 3 and cells[0] and not cells[0].startswith('-'):
                    endpoint = cells[0].strip()
                    method = cells[1].strip()
                    opensearch_apis_str = cells[2].strip()

                    # Skip header rows
                    if 'OSD Endpoint' in endpoint:
                        continue

                    # Parse OpenSearch API calls
                    # Remove markdown formatting
                    opensearch_apis_str = re.sub(r'`', '', opensearch_apis_str)

                    # Split by common separators
                    api_calls = re.split(r'[,+]|\s+\+\s+|,\s+', opensearch_apis_str)

                    for api_call in api_calls:
                        api_call = api_call.strip()
                        if api_call and api_call not in ['(none)', '(any)', '(reads cached status)']:
                            # Clean up the API name
                            api_call = re.sub(r'\(.*?\)|\[.*?\]', '', api_call).strip()

                            if api_call:
                                if api_call not in api_usage_map:
                                    api_usage_map[api_call] = []

                                usage_info = f"{plugin_name}: {endpoint} ({method})"
                                if usage_info not in api_usage_map[api_call]:
                                    api_usage_map[api_call].append(usage_info)

    return api_usage_map

def categorize_opensearch_api(api_name):
    """Categorize OpenSearch API into groups"""
    api_lower = api_name.lower()

    if any(x in api_lower for x in ['search', 'msearch', 'scroll', 'mget', 'count']):
        return 'Document - Search & Retrieval'
    elif any(x in api_lower for x in ['index', 'create', 'bulk', 'update', 'delete', 'reindex']):
        return 'Document - Write & Modify'
    elif api_lower.startswith('indices.'):
        return 'Index Management'
    elif api_lower.startswith('cluster.'):
        return 'Cluster APIs'
    elif api_lower.startswith('nodes.'):
        return 'Node APIs'
    elif api_lower.startswith('cat.'):
        return 'Cat APIs'
    elif api_lower.startswith('tasks.'):
        return 'Task Management'
    elif 'transport.request' in api_lower:
        return 'Custom/Plugin APIs'
    elif api_lower in ['info', 'ping']:
        return 'Cluster Info'
    else:
        return 'Other'

def create_opensearch_api_excel(api_usage_map, output_file):
    """Create Excel spreadsheet organized by OpenSearch API"""
    wb = Workbook()
    ws = wb.active
    ws.title = "OpenSearch API Usage"

    # Define headers
    headers = [
        'API Category',
        'OpenSearch API',
        'Used By OSD Features',
        '# of Usages',
        'Team Assignment',
        'SDM Owner',
        'Priority',
        'Status',
        'Notes'
    ]

    # Style definitions
    header_fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Write headers
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = border

    # Set column widths
    column_widths = {
        'A': 25,  # API Category
        'B': 30,  # OpenSearch API
        'C': 60,  # Used By OSD Features
        'D': 12,  # # of Usages
        'E': 20,  # Team Assignment
        'F': 20,  # SDM Owner
        'G': 12,  # Priority
        'H': 15,  # Status
        'I': 40   # Notes
    }

    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width

    # Color coding for different API categories
    category_colors = {
        'Document - Search & Retrieval': 'B4C7E7',
        'Document - Write & Modify': 'C5E0B4',
        'Index Management': 'FFE699',
        'Cluster APIs': 'F4B183',
        'Node APIs': 'F8CBAD',
        'Cat APIs': 'E2EFDA',
        'Task Management': 'FCE4D6',
        'Custom/Plugin APIs': 'FFF2CC',
        'Cluster Info': 'DDEBF7',
        'Other': 'F2F2F2'
    }

    # Prepare data - categorize and sort
    categorized_data = []
    for api_name, usages in sorted(api_usage_map.items()):
        category = categorize_opensearch_api(api_name)
        categorized_data.append({
            'category': category,
            'api_name': api_name,
            'usages': usages,
            'usage_count': len(usages)
        })

    # Sort by category then by API name
    categorized_data.sort(key=lambda x: (x['category'], x['api_name']))

    # Write data rows
    row_num = 2
    for data in categorized_data:
        category_color = category_colors.get(data['category'], category_colors['Other'])
        row_fill = PatternFill(start_color=category_color, end_color=category_color, fill_type="solid")

        # Join usages with line breaks
        usages_text = '\n'.join(data['usages']) if data['usages'] else 'No direct usage found'

        row_data = [
            data['category'],
            data['api_name'],
            usages_text,
            data['usage_count'],
            '',  # Team Assignment
            '',  # SDM Owner
            '',  # Priority
            '',  # Status
            ''   # Notes
        ]

        for col_num, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.value = value
            cell.border = border
            cell.alignment = Alignment(vertical='top', wrap_text=True)

            # Apply color to first 4 columns (data columns)
            if col_num <= 4:
                cell.fill = row_fill

        # Set row height to accommodate multiple usages
        if data['usage_count'] > 3:
            ws.row_dimensions[row_num].height = min(15 * data['usage_count'], 200)

        row_num += 1

    # Freeze panes (freeze header row)
    ws.freeze_panes = 'A2'

    # Add auto-filter
    ws.auto_filter.ref = f"A1:I{row_num - 1}"

    # Create summary sheet
    ws_summary = wb.create_sheet("Summary", 0)
    ws_summary['A1'] = "OpenSearch API Usage Summary"
    ws_summary['A1'].font = Font(bold=True, size=14)

    ws_summary['A3'] = "Total OpenSearch APIs:"
    ws_summary['B3'] = len(api_usage_map)
    ws_summary['B3'].font = Font(bold=True)

    ws_summary['A5'] = "Breakdown by Category:"
    ws_summary['A5'].font = Font(bold=True)

    # Count by category
    category_counts = {}
    for data in categorized_data:
        category = data['category']
        category_counts[category] = category_counts.get(category, 0) + 1

    row = 6
    for category, count in sorted(category_counts.items(), key=lambda x: -x[1]):
        ws_summary[f'A{row}'] = category
        ws_summary[f'B{row}'] = count

        # Apply category color
        category_color = category_colors.get(category, category_colors['Other'])
        fill = PatternFill(start_color=category_color, end_color=category_color, fill_type="solid")
        ws_summary[f'A{row}'].fill = fill

        row += 1

    ws_summary['A' + str(row + 1)] = "Most Used APIs:"
    ws_summary['A' + str(row + 1)].font = Font(bold=True)

    # Get top 10 most used APIs
    top_apis = sorted(categorized_data, key=lambda x: x['usage_count'], reverse=True)[:10]
    row += 2
    for i, data in enumerate(top_apis, 1):
        ws_summary[f'A{row}'] = data['api_name']
        ws_summary[f'B{row}'] = data['usage_count']
        row += 1

    ws_summary.column_dimensions['A'].width = 35
    ws_summary.column_dimensions['B'].width = 15

    # Create API Groups sheet
    ws_groups = wb.create_sheet("API Groups", 2)
    ws_groups['A1'] = "OpenSearch API Groups & Descriptions"
    ws_groups['A1'].font = Font(bold=True, size=14)

    api_groups = [
        ("Document - Search & Retrieval", "search, msearch, mget, scroll, count", "Query and retrieve documents from indices"),
        ("Document - Write & Modify", "index, create, bulk, update, delete, reindex, updateByQuery, deleteByQuery", "Create, update, and delete documents"),
        ("Index Management", "indices.create, indices.delete, indices.get, indices.exists, indices.updateAliases, etc.", "Manage index lifecycle, mappings, settings, and aliases"),
        ("Cluster APIs", "cluster.state, cluster.getSettings, cluster.stats", "Retrieve cluster-level information and settings"),
        ("Node APIs", "nodes.info, nodes.usage", "Get information about cluster nodes"),
        ("Cat APIs", "cat.indices, cat.plugins", "Compact text output for human consumption"),
        ("Task Management", "tasks.get, tasks.cancel", "Manage long-running tasks"),
        ("Custom/Plugin APIs", "transport.request (ML, PPL, etc.)", "Direct access to custom OpenSearch plugin APIs"),
    ]

    row = 3
    ws_groups['A' + str(row)] = "Category"
    ws_groups['B' + str(row)] = "APIs"
    ws_groups['C' + str(row)] = "Description"

    for col in ['A', 'B', 'C']:
        ws_groups[col + str(row)].font = Font(bold=True)
        ws_groups[col + str(row)].fill = header_fill
        ws_groups[col + str(row)].font = Font(bold=True, color="FFFFFF")

    row += 1
    for group_name, apis, description in api_groups:
        ws_groups[f'A{row}'] = group_name
        ws_groups[f'B{row}'] = apis
        ws_groups[f'C{row}'] = description
        ws_groups[f'A{row}'].alignment = Alignment(vertical='top', wrap_text=True)
        ws_groups[f'B{row}'].alignment = Alignment(vertical='top', wrap_text=True)
        ws_groups[f'C{row}'].alignment = Alignment(vertical='top', wrap_text=True)
        row += 1

    ws_groups.column_dimensions['A'].width = 30
    ws_groups.column_dimensions['B'].width = 50
    ws_groups.column_dimensions['C'].width = 40

    # Save workbook
    wb.save(output_file)
    print(f"✓ Excel spreadsheet created: {output_file}")
    print(f"  - Total OpenSearch APIs: {len(api_usage_map)}")
    print(f"  - API Categories: {len(category_counts)}")
    print(f"  - Sheets: Summary, OpenSearch API Usage, API Groups")
    print(f"  - Columns: Team Assignment, SDM Owner, Priority, Status, Notes ready for input")

if __name__ == '__main__':
    input_file = 'OSD_TO_OPENSEARCH_API_MAPPING.md'
    output_file = 'mustang/OpenSearch_API_Usage_TeamAssignment.xlsx'

    print(f"Parsing {input_file}...")
    api_usage_map = parse_opensearch_api_usage(input_file)

    print(f"Creating Excel spreadsheet...")
    create_opensearch_api_excel(api_usage_map, output_file)

    print("\n✓ Done! Open the Excel file to assign teams and SDM owners to OpenSearch APIs.")
