#!/usr/bin/env python3
"""
Generate Excel spreadsheet from OSD-to-OpenSearch API mapping
with columns for team and SDM owner assignment
"""

import re
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def parse_mapping_markdown(file_path):
    """Parse the markdown file and extract API mappings"""
    with open(file_path, 'r', encoding='utf-8') as f:
        content = f.read()

    mappings = []
    current_section = None

    # Find all sections with endpoint mappings
    sections = [
        ('Core: Saved Objects API', 'Core'),
        ('Core: Status & Health', 'Core'),
        ('Plugin: Data (Search)', 'Data'),
        ('Plugin: Console (Dev Tools Proxy)', 'Console'),
        ('Plugin: Data Importer', 'Data Importer'),
        ('Plugin: Application Config', 'Application Config'),
        ('Plugin: Region Map', 'Region Map'),
        ('Plugin: Index Pattern Management', 'Index Pattern Management'),
        ('Plugin: Query Enhancements', 'Query Enhancements'),
        ('Plugin: Data Source Management', 'Data Source Management'),
        ('Plugin: Chat (ML/AI)', 'Chat/ML'),
        ('Plugin: Workspace', 'Workspace'),
        ('Plugin: Home (Sample Data)', 'Home'),
        ('Plugin: Telemetry', 'Telemetry'),
        ('Plugin: Vis Type Timeline', 'Timeline'),
        ('Plugin: Vis Type Timeseries (TSVB)', 'TSVB'),
        ('Plugin: Data Source', 'Data Source'),
        ('Plugin: Saved Objects Management', 'Saved Objects Management'),
        ('Plugin: Share (Short URLs)', 'Share'),
        ('Plugin: Legacy Export', 'Legacy Export'),
        ('Core: Dynamic Config Store (Internal)', 'Core (Internal)'),
        ('Core: Saved Objects Migrations (Internal)', 'Core (Internal)'),
    ]

    for section_header, plugin_name in sections:
        # Find the section in the markdown
        pattern = rf'## {re.escape(section_header)}.*?\n(.*?)(?=\n## |\Z)'
        match = re.search(pattern, content, re.DOTALL)

        if not match:
            continue

        section_content = match.group(1)

        # Look for endpoint mapping tables
        # Pattern: | OSD Endpoint | Method | OpenSearch APIs Called | Data Flow |
        table_pattern = r'\| OSD Endpoint \| Method \| OpenSearch APIs Called \| Data Flow \|.*?\n\|.*?\n((?:\|.*?\n)*)'
        table_match = re.search(table_pattern, section_content, re.DOTALL)

        if table_match:
            table_rows = table_match.group(1).strip().split('\n')
            for row in table_rows:
                # Parse table row
                cells = [cell.strip() for cell in row.split('|')[1:-1]]  # Remove empty first/last
                if len(cells) >= 4 and cells[0] and not cells[0].startswith('-'):
                    endpoint = cells[0].strip()
                    method = cells[1].strip()
                    opensearch_apis = cells[2].strip()
                    data_flow = cells[3].strip() if len(cells) > 3 else ''

                    # Skip header rows
                    if 'OSD Endpoint' in endpoint or '---' in endpoint:
                        continue

                    mappings.append({
                        'plugin': plugin_name,
                        'endpoint': endpoint,
                        'method': method,
                        'opensearch_apis': opensearch_apis,
                        'data_flow': data_flow,
                        'team': '',
                        'sdm_owner': '',
                        'priority': '',
                        'status': '',
                        'notes': ''
                    })

    return mappings

def create_excel_spreadsheet(mappings, output_file):
    """Create Excel spreadsheet with API mappings"""
    wb = Workbook()
    ws = wb.active
    ws.title = "OSD API Mappings"

    # Define headers
    headers = [
        'Plugin/Area',
        'OSD API Endpoint',
        'HTTP Method',
        'OpenSearch APIs Used',
        'Description/Data Flow',
        'Team Assignment',
        'SDM Owner',
        'Priority',
        'Status',
        'Notes'
    ]

    # Style definitions
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Write headers
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = border

    # Set column widths
    column_widths = {
        'A': 20,  # Plugin/Area
        'B': 40,  # OSD API Endpoint
        'C': 12,  # HTTP Method
        'D': 35,  # OpenSearch APIs Used
        'E': 50,  # Description/Data Flow
        'F': 20,  # Team Assignment
        'G': 20,  # SDM Owner
        'H': 12,  # Priority
        'I': 15,  # Status
        'J': 40   # Notes
    }

    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width

    # Color coding for different plugins
    plugin_colors = {
        'Core': 'E7E6E6',
        'Core (Internal)': 'D9D9D9',
        'Data': 'B4C7E7',
        'Console': 'C5E0B4',
        'Data Importer': 'FFD966',
        'Application Config': 'F4B183',
        'Query Enhancements': 'C6E0B4',
        'Data Source Management': 'A9D08E',
        'Chat/ML': 'FFE699',
        'Workspace': 'BDD7EE',
        'Default': 'FFFFFF'
    }

    # Write data rows
    for row_num, mapping in enumerate(mappings, 2):
        plugin_color = plugin_colors.get(mapping['plugin'], plugin_colors['Default'])
        row_fill = PatternFill(start_color=plugin_color, end_color=plugin_color, fill_type="solid")

        row_data = [
            mapping['plugin'],
            mapping['endpoint'],
            mapping['method'],
            mapping['opensearch_apis'],
            mapping['data_flow'],
            mapping['team'],
            mapping['sdm_owner'],
            mapping['priority'],
            mapping['status'],
            mapping['notes']
        ]

        for col_num, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.value = value
            cell.border = border
            cell.alignment = Alignment(vertical='top', wrap_text=True)

            # Apply color to first 5 columns (data columns)
            if col_num <= 5:
                cell.fill = row_fill

    # Freeze panes (freeze header row)
    ws.freeze_panes = 'A2'

    # Add auto-filter
    ws.auto_filter.ref = f"A1:J{len(mappings) + 1}"

    # Create summary sheet
    ws_summary = wb.create_sheet("Summary", 0)
    ws_summary['A1'] = "OSD API to OpenSearch API Mapping Summary"
    ws_summary['A1'].font = Font(bold=True, size=14)
    ws_summary['A3'] = "Total API Endpoints:"
    ws_summary['B3'] = len(mappings)
    ws_summary['B3'].font = Font(bold=True)

    ws_summary['A5'] = "Breakdown by Plugin/Area:"
    ws_summary['A5'].font = Font(bold=True)

    # Count by plugin
    plugin_counts = {}
    for mapping in mappings:
        plugin = mapping['plugin']
        plugin_counts[plugin] = plugin_counts.get(plugin, 0) + 1

    row = 6
    for plugin, count in sorted(plugin_counts.items(), key=lambda x: -x[1]):
        ws_summary[f'A{row}'] = plugin
        ws_summary[f'B{row}'] = count
        row += 1

    ws_summary.column_dimensions['A'].width = 30
    ws_summary.column_dimensions['B'].width = 15

    # Save workbook
    wb.save(output_file)
    print(f"✓ Excel spreadsheet created: {output_file}")
    print(f"  - Total endpoints: {len(mappings)}")
    print(f"  - Plugins/areas: {len(plugin_counts)}")
    print(f"  - Columns: Team Assignment, SDM Owner, Priority, Status, Notes ready for input")

if __name__ == '__main__':
    import sys

    input_file = 'OSD_TO_OPENSEARCH_API_MAPPING.md'
    output_file = 'mustang/OSD_API_Mapping_TeamAssignment.xlsx'

    print(f"Parsing {input_file}...")
    mappings = parse_mapping_markdown(input_file)

    print(f"Creating Excel spreadsheet...")
    create_excel_spreadsheet(mappings, output_file)

    print("\n✓ Done! Open the Excel file to assign teams and SDM owners.")
